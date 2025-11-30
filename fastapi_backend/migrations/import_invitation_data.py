#!/usr/bin/env python3
"""
Import invitation data from Excel files into MongoDB invitations collection
Creates invitations in 'pending' status ready for bulk sending via UI

Usage:
    python import_invitation_data.py --dry-run  # Preview only
    python import_invitation_data.py            # Import data
"""

import asyncio
import sys
import os
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
import secrets
import string

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from motor.motor_asyncio import AsyncIOMotorClient
from config import Settings

settings = Settings()


def generate_token(length: int = 32) -> str:
    """Generate secure random token for invitation"""
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))


async def import_invitations(dry_run=True):
    """Import invitation data from Excel files"""
    
    client = AsyncIOMotorClient(settings.mongodb_url)
    db = client[settings.database_name]
    
    print("=" * 80)
    print("INVITATION DATA IMPORT")
    print("=" * 80)
    print(f"Mode: {'🔍 DRY RUN (no changes)' if dry_run else '⚡ LIVE (importing data)'}")
    print(f"Database: {settings.database_name}")
    print(f"Timestamp: {datetime.utcnow().isoformat()}")
    print("=" * 80)
    print()
    
    # Data files
    data_dir = Path(__file__).parent.parent.parent / "data_migration"
    files = [
        data_dir / "girls_list.xlsx",
        data_dir / "boys_List.xlsx"
    ]
    
    all_invitations = []
    stats = {
        'total_rows': 0,
        'valid_emails': 0,
        'invalid_emails': 0,
        'duplicates': 0,
        'to_import': 0
    }
    
    # Get existing emails to check for duplicates
    existing_emails = set()
    if not dry_run:
        cursor = db.invitations.find({}, {"email": 1})
        async for doc in cursor:
            if doc.get("email"):
                existing_emails.add(doc["email"].lower())
    
    print(f"📊 Found {len(existing_emails)} existing invitations in database")
    print()
    
    # Process each file
    for file_path in files:
        if not file_path.exists():
            print(f"⚠️  File not found: {file_path}")
            continue
            
        print(f"📂 Processing: {file_path.name}")
        print("-" * 80)
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"   Columns: {headers}")
        
        # Process rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            stats['total_rows'] += 1
            
            # Parse row
            gender = row[0] if len(row) > 0 else None
            name = row[1] if len(row) > 1 else None
            phone = row[2] if len(row) > 2 else None
            email = row[3] if len(row) > 3 else None
            
            # Validate email
            if not email or '@' not in str(email):
                stats['invalid_emails'] += 1
                print(f"   ⚠️  Row {row_idx}: Invalid email: {email}")
                continue
            
            email = str(email).strip().lower()
            
            # Check for duplicates
            if email in existing_emails or email in [inv['email'] for inv in all_invitations]:
                stats['duplicates'] += 1
                continue
            
            stats['valid_emails'] += 1
            
            # Normalize gender for comments
            gender_comment = "Female" if gender and "girl" in str(gender).lower() else "Male"
            
            # Generate invitation token (CRITICAL: needed for registration tracking)
            token = generate_token()
            token_expiry = datetime.utcnow() + timedelta(days=30)  # 30-day expiry
            
            # Create invitation record
            invitation = {
                "name": str(name).strip() if name else "Unknown",
                "email": email,
                "phone": str(phone).strip() if phone else "",
                "invitedBy": "admin",
                "channel": "email",
                "emailStatus": "pending",  # Not sent yet
                "smsStatus": "not_sent",
                "customMessage": "",
                "emailSubject": "You're Invited to Join USVedika for US Citizens & GC Holders",
                "comments": gender_comment,  # Store gender in comments
                "createdAt": datetime.utcnow(),
                "updatedAt": datetime.utcnow(),
                "archived": False,
                # Invitation token (CRITICAL for tracking registrations)
                "invitationToken": token,
                "tokenExpiresAt": token_expiry
            }
            
            all_invitations.append(invitation)
        
        print(f"   ✅ Processed {stats['total_rows']} rows from {file_path.name}")
        print()
    
    stats['to_import'] = len(all_invitations)
    
    # Summary
    print("=" * 80)
    print("IMPORT SUMMARY")
    print("=" * 80)
    print(f"Total rows processed: {stats['total_rows']}")
    print(f"✅ Valid emails: {stats['valid_emails']}")
    print(f"❌ Invalid emails: {stats['invalid_emails']}")
    print(f"🔄 Duplicates skipped: {stats['duplicates']}")
    print(f"📥 Ready to import: {stats['to_import']}")
    print("=" * 80)
    print()
    
    # Import data
    if dry_run:
        print("💡 This was a DRY RUN - no changes were made")
        print("   Run without --dry-run to import data")
        print()
        
        # Show sample records
        if all_invitations:
            print("📋 Sample records (first 5):")
            for inv in all_invitations[:5]:
                print(f"   - {inv['name']} ({inv['comments']}) <{inv['email']}>")
                print(f"     Subject: {inv['emailSubject']}")
                print(f"     Token: {inv['invitationToken'][:16]}... (expires: {inv['tokenExpiresAt'].strftime('%Y-%m-%d')})")
            print()
    else:
        if all_invitations:
            print("⚡ Importing invitations to database...")
            result = await db.invitations.insert_many(all_invitations)
            print(f"✅ Successfully imported {len(result.inserted_ids)} invitations!")
            print()
            print("🎯 Next steps:")
            print("   1. Go to https://l3v3lmatches.com/invitations")
            print("   2. Select invitations to send")
            print("   3. Click 'Send Selected' button")
            print()
        else:
            print("⚠️  No new invitations to import")
            print()
    
    client.close()


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Import invitation data from Excel files')
    parser.add_argument('--dry-run', action='store_true', help='Preview without making changes')
    args = parser.parse_args()
    
    asyncio.run(import_invitations(dry_run=args.dry_run))


if __name__ == "__main__":
    main()
